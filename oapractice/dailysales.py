import openpyxl
import codecs
import pandas as pd
from openpyxl.utils import get_column_letter

'''
# 将text文件转化为xlsx
'''
my_asins = ['B01N76O2E9','B079P1X8H8','B07J4ZQDY4','B07ZTDFKLK','B078B7WT3R','B071H2YTQ9','B071H2YTQ9','B07QCSGQPV','B079M3W46N','B079MHVMKJ','B079LZXDWV','B07D32FNDZ','B07BN8JZJP','B07MVZMM7Z','B07DZS6J5P','B07DZR15FG','B07JB9NF24','B07VC8NN92','B07VJNQV4T','B07W3M4VGS','B07YV8LYZN','B07W3LF4HR','B07YV8LYZN',]

def txt_to_xlsx(filename,outfile): 
    fr = codecs.open(filename,'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    # ws = wb.create_sheet()
    ws.title = 'orders'
    
    row = 0
    for line in fr:
        row +=1
        line = line.strip()
        line = line.split('\t')
        
        col = 0
        for j in range(len(line)):
            col +=1
            # print (line[j])
            ws.cell(column = col, row = row, value = line[j].format(get_column_letter(col)))
    
    wb.save(outfile)


def cal_sales(filename):
    wb = pd.read_excel(filename)
    all_sales = wb.iloc[:,[12,15,17]].groupby('asin').sum()
    all_sales = all_sales.reset_index()
    my_sales = all_sales[all_sales['asin'].isin(my_asins)]

    print(my_sales)
    print('==========')
    print(my_sales['quantity'].sum())
    print(my_sales['item-price'].sum())


'''
# #读取xlsx内容    
# def read_xlsx(filename):
#     #载入文件
#     wb = openpyxl.load_workbook(filename)
#     #获取Sheet1工作表
#     ws = wb.get_sheet_by_name('Sheet1')
#     #按行读取
#     for row in ws.rows:
#         for cell in row:
#             print (cell.value)
#     #按列读
#     for col in ws.columns:
#         for cell in col:
#             print (cell.value)
'''

if __name__=='__main__':
    inputfileTxt = 'orders.txt'
    outfileExcel = 'daily_sales.xlsx'
    txt_to_xlsx(inputfileTxt,outfileExcel)

    cal_sales(outfileExcel)
    # read_xlsx(outfileExcel)